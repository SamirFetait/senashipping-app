"""
Excel report generation for loading conditions.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from ..config.limits import MASS_PER_HEAD_T
from ..repositories import database
from ..repositories.livestock_pen_repository import LivestockPenRepository
from ..repositories.tank_repository import TankRepository

if TYPE_CHECKING:
    from ..models import Ship, Voyage, LoadingCondition
    from ..services.stability_service import ConditionResults


def _fmt(value: object, fmt: str) -> str:
    """Safely format numeric values, falling back to string/blank."""
    if value is None:
        return ""
    try:
        return format(float(value), fmt)
    except (TypeError, ValueError):
        return str(value)


def _style_header(ws) -> None:
    """Apply a simple header style similar to the manual tables."""
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def _style_body_table(ws, *, start_row: int = 2, first_col_bold: bool = True, stripe: bool = True) -> None:
    """
    Apply professional-looking body styling:
    - Optional zebra striping
    - Bold first column
    - Left-align text in first column, right-align numeric-looking cells elsewhere
    """
    stripe_fill = PatternFill(fill_type="solid", fgColor="F5F5F5")
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if first_col_bold and row and row[0].value not in (None, ""):
            row[0].font = Font(bold=True)
        for cell in row:
            # Alternate row shading (based on Excel row index)
            if stripe and cell.row % 2 == 0:
                # Don't overwrite header styling
                if cell.fill is None or cell.fill.fill_type is None:
                    cell.fill = stripe_fill
            # Alignment: first column left, others right-ish
            if cell.column == 1:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True,
                )


def _deck_to_letter(deck: str) -> str | None:
    """Normalize deck value to A–H so it matches loading condition decks."""
    s = (deck or "").strip().upper()
    if not s:
        return None
    deck_letters = ("A", "B", "C", "D", "E", "F", "G", "H")
    if s in deck_letters:
        return s
    if s.isdigit() and 1 <= int(s) <= 8:
        return chr(ord("A") + int(s) - 1)
    if s.startswith("DK") and s[2:].strip().isdigit():
        n = int(s[2:].strip())
        if 1 <= n <= 8:
            return chr(ord("A") + n - 1)
    return s if s in deck_letters else None


def _build_weight_items_rows(ship, condition, results) -> list[dict] | None:
    """
    Build per-item weight / FSM-style rows for Excel, mirroring the PDF:

    Item | Quantity / Fill | Unit mass (t) | Total mass (t) |
    Long. arm (m) | Vert. arm (m) | Total FSM (t·m) | FSM Type
    """
    pen_loadings = getattr(condition, "pen_loadings", None) or {}
    tank_volumes = getattr(condition, "tank_volumes_m3", None) or {}
    if not pen_loadings and not tank_volumes:
        return None

    pens = []
    tanks = []
    if getattr(ship, "id", None) and database.SessionLocal is not None:
        with database.SessionLocal() as db:
            pens = LivestockPenRepository(db).list_for_ship(ship.id)
            tanks = TankRepository(db).list_for_ship(ship.id)

    rows: list[dict] = []

    # --- Pens grouped by deck (DECK A, DECK B, ...) ---
    if pens and pen_loadings:
        pen_by_id = {p.id: p for p in pens if p.id is not None}
        deck_groups: dict[str, dict[str, float]] = {}
        for pen_id, heads in pen_loadings.items():
            try:
                heads_int = int(heads)
            except (TypeError, ValueError):
                continue
            if heads_int <= 0:
                continue
            pen = pen_by_id.get(pen_id)
            if not pen:
                continue
            deck_letter = _deck_to_letter(getattr(pen, "deck", "") or "") or (getattr(pen, "deck", "") or "")
            if not deck_letter:
                continue
            grp = deck_groups.setdefault(
                deck_letter,
                {"heads": 0.0, "mass": 0.0, "lcg_moment": 0.0, "vcg_moment": 0.0},
            )
            mass = heads_int * MASS_PER_HEAD_T
            grp["heads"] += heads_int
            grp["mass"] += mass
            grp["lcg_moment"] += mass * float(getattr(pen, "lcg_m", 0.0) or 0.0)
            grp["vcg_moment"] += mass * float(getattr(pen, "vcg_m", 0.0) or 0.0)

        for deck_key in sorted(deck_groups):
            grp = deck_groups[deck_key]
            mass = grp["mass"]
            heads_total = int(grp["heads"])
            if heads_total <= 0:
                continue
            long_arm = grp["lcg_moment"] / mass if mass > 0 else 0.0
            vert_arm = grp["vcg_moment"] / mass if mass > 0 else 0.0
            rows.append(
                {
                    "Item": f"DECK {deck_key}",
                    "Quantity / Fill": heads_total,
                    "Unit mass (t)": float(f"{MASS_PER_HEAD_T:.4f}"),
                    "Total mass (t)": float(f"{mass:.4f}"),
                    "Long. arm (m)": float(f"{long_arm:.4f}") if mass > 0 else "",
                    "Vert. arm (m)": float(f"{vert_arm:.4f}") if mass > 0 else "",
                    "Total FSM (t·m)": "",
                    "FSM Type": "N/A (pens)",
                }
            )
    elif pen_loadings:
        # Fallback without DB: aggregate all pens into one row.
        total_heads = 0
        for h in pen_loadings.values():
            try:
                total_heads += max(0, int(h))
            except (TypeError, ValueError):
                continue
        if total_heads > 0:
            mass = total_heads * MASS_PER_HEAD_T
            rows.append(
                {
                    "Item": "Livestock (pens)",
                    "Quantity / Fill": total_heads,
                    "Unit mass (t)": float(f"{MASS_PER_HEAD_T:.4f}"),
                    "Total mass (t)": float(f"{mass:.4f}"),
                    "Long. arm (m)": "",
                    "Vert. arm (m)": "",
                    "Total FSM (t·m)": "",
                    "FSM Type": "N/A (pens)",
                }
            )

    # --- Tanks: one row per tank with volume ---
    snapshot = getattr(results, "snapshot", None)
    inputs = getattr(snapshot, "inputs", {}) if snapshot else {}
    try:
        cargo_density = float(inputs.get("cargo_density_t_per_m3", 1.0) or 1.0)
    except (TypeError, ValueError):
        cargo_density = 1.0

    total_tank_mass = 0.0
    if tanks and tank_volumes:
        tank_by_id = {t.id: t for t in tanks if t.id is not None}
        L = float(getattr(ship, "length_overall_m", 0.0) or 0.0)
        if L <= 0.0:
            L = 0.0
        for tid, vol_m3 in tank_volumes.items():
            try:
                vol = float(vol_m3)
            except (TypeError, ValueError):
                continue
            if vol <= 0.0:
                continue
            tank = tank_by_id.get(tid)
            name = getattr(tank, "name", None) if tank else None
            item_label = name or f"Tank {tid}"
            mass_t = vol * cargo_density
            total_tank_mass += mass_t

            if tank and L > 0.0:
                pos = float(getattr(tank, "longitudinal_pos", 0.0) or 0.0)
                if pos > 1.5:
                    lcg_m = pos
                else:
                    lcg_m = pos * L
            else:
                lcg_m = 0.0
            vcg_m = float(getattr(tank, "kg_m", 0.0) or 0.0) if tank else 0.0

            rows.append(
                {
                    "Item": item_label,
                    "Quantity / Fill": float(f"{vol:.4f}"),
                    "Unit mass (t)": float(f"{cargo_density:.4f}"),
                    "Total mass (t)": float(f"{mass_t:.4f}"),
                    "Long. arm (m)": float(f"{lcg_m:.4f}") if lcg_m else "",
                    "Vert. arm (m)": float(f"{vcg_m:.4f}") if vcg_m else "",
                    "Total FSM (t·m)": "",
                    "FSM Type": "N/A (tanks)",
                }
            )

    # --- Aggregate FSM row (all tanks) ---
    if tank_volumes:
        gm_raw = getattr(results, "gm_m", None)
        validation = getattr(results, "validation", None)
        gm_eff = getattr(validation, "gm_effective", None) if validation else None
        total_fsm = ""
        fsm_type = "N/A"
        try:
            if gm_raw is not None and gm_eff is not None and float(gm_raw) > 0:
                fsc = float(gm_raw) - float(gm_eff)
                if fsc > 0 and getattr(results, "displacement_t", None):
                    total_fsm_val = fsc * float(results.displacement_t)
                    total_fsm = float(f"{total_fsm_val:.4f}")
                    fsm_type = "Aggregate FSM (from GM eff.)"
        except (TypeError, ValueError):
            total_fsm = ""
            fsm_type = "N/A"

        rows.append(
            {
                "Item": "FSM total (all tanks)",
                "Quantity / Fill": "",
                "Unit mass (t)": "",
                "Total mass (t)": "",
                "Long. arm (m)": "",
                "Vert. arm (m)": "",
                "Total FSM (t·m)": total_fsm,
                "FSM Type": fsm_type,
            }
        )

    return rows if rows else None
            else:
                cell.alignment = Alignment(
                    horizontal="right",
                    vertical="center",
                    wrap_text=True,
                )


def export_condition_to_excel(
    filepath: Path,
    ship: "Ship",
    voyage: "Voyage",
    condition: "LoadingCondition",
    results: "ConditionResults",
) -> None:
    """
    Generate a multi-sheet Excel report for a loading condition.

    The layout is inspired by the vessel loading manual pages:
    - Condition summary
    - Equilibrium / hydrostatic-style data
    - IMO / ancillary criteria with pass/fail highlighting
    """
    # --- Sheet 1: high-level condition summary ---
    summary_data = {
        "Parameter": [
            "Ship",
            "IMO",
            "Voyage",
            "Departure",
            "Arrival",
            "Condition",
            "Displacement (t)",
            "Draft mid (m)",
            "Draft aft (m)",
            "Draft fwd (m)",
            "Trim (m, +ve stern down)",
            "Heel (deg)",
            "GM (effective, m)",
            "GM (raw, m)",
            "KG (m)",
            "KM (m)",
        ],
        "Value": [
            ship.name,
            getattr(ship, "imo_number", "") or "",
            voyage.name,
            voyage.departure_port,
            voyage.arrival_port,
            condition.name,
            _fmt(getattr(condition, "displacement_t", None), ".1f"),
            _fmt(getattr(condition, "draft_m", None), ".3f"),
            _fmt(getattr(results, "draft_aft_m", None), ".3f"),
            _fmt(getattr(results, "draft_fwd_m", None), ".3f"),
            _fmt(getattr(condition, "trim_m", None), ".3f"),
            _fmt(getattr(results, "heel_deg", None), ".2f"),
            _fmt(getattr(getattr(results, "validation", None), "gm_effective", None), ".3f"),
            _fmt(getattr(results, "gm_m", None), ".3f"),
            _fmt(getattr(results, "kg_m", None), ".3f"),
            _fmt(getattr(results, "km_m", None), ".3f"),
        ],
    }
    strength = getattr(results, "strength", None)
    if strength:
        summary_data["Parameter"].append("SWBM approx. (tm)")
        summary_data["Value"].append(_fmt(getattr(strength, "still_water_bm_approx_tm", None), ".0f"))

    ancillary = getattr(results, "ancillary", None)
    if ancillary:
        summary_data["Parameter"].extend(
            [
                "Propeller immersion (%)",
                "Visibility ahead (m)",
                "Air draft (m)",
                "GZ criteria OK",
            ]
        )
        summary_data["Value"].extend(
            [
                _fmt(getattr(ancillary, "prop_immersion_pct", None), ".1f"),
                _fmt(getattr(ancillary, "visibility_m", None), ".1f"),
                _fmt(getattr(ancillary, "air_draft_m", None), ".2f"),
                "YES" if getattr(ancillary, "gz_criteria_ok", False) else "NO",
            ]
        )

    df_summary = pd.DataFrame(summary_data)

    # --- Sheet 2: equilibrium-style data (narrower, with units) ---
    eq_rows = {
        "Parameter": [
            "Displacement",
            "Draft amidships",
            "Draft at AP",
            "Draft at FP",
            "Trim (stern down +ve)",
            "Heel",
            "GM (effective)",
            "GM (raw)",
            "KG",
            "KM",
        ],
        "Value": [
            _fmt(getattr(results, "displacement_t", None), ".1f"),
            _fmt(getattr(results, "draft_m", None), ".3f"),
            _fmt(getattr(results, "draft_aft_m", None), ".3f"),
            _fmt(getattr(results, "draft_fwd_m", None), ".3f"),
            _fmt(getattr(results, "trim_m", None), ".3f"),
            _fmt(getattr(results, "heel_deg", None), ".2f"),
            _fmt(getattr(getattr(results, "validation", None), "gm_effective", None), ".3f"),
            _fmt(getattr(results, "gm_m", None), ".3f"),
            _fmt(getattr(results, "kg_m", None), ".3f"),
            _fmt(getattr(results, "km_m", None), ".3f"),
        ],
        "Unit": [
            "t",
            "m",
            "m",
            "m",
            "m",
            "deg",
            "m",
            "m",
            "m",
            "m",
        ],
    }
    if strength:
        eq_rows["Parameter"].append("SWBM approx.")
        eq_rows["Value"].append(_fmt(getattr(strength, "still_water_bm_approx_tm", None), ".0f"))
        eq_rows["Unit"].append("tm")

    if ancillary:
        eq_rows["Parameter"].extend(
            [
                "Propeller immersion",
                "Visibility ahead",
                "Air draft",
            ]
        )
        eq_rows["Value"].extend(
            [
                _fmt(getattr(ancillary, "prop_immersion_pct", None), ".1f"),
                _fmt(getattr(ancillary, "visibility_m", None), ".1f"),
                _fmt(getattr(ancillary, "air_draft_m", None), ".2f"),
            ]
        )
        eq_rows["Unit"].extend(["% dia", "m", "m"])

    df_eq = pd.DataFrame(eq_rows)

    # --- Sheet 3: IMO / ancillary criteria table, if available ---
    criteria = getattr(results, "criteria", None)
    crit_df = None
    if criteria is not None and getattr(criteria, "lines", None):
        crit_rows = []
        for line in criteria.lines:
            result_obj = getattr(line, "result", None)
            result_str = getattr(result_obj, "name", str(result_obj)) if result_obj is not None else ""
            crit_rows.append(
                {
                    "Group": getattr(line, "parent_code", "") or "",
                    "Code": getattr(line, "code", "") or "",
                    "Name": getattr(line, "name", "") or "",
                    "Reference": getattr(line, "reference", "") or "",
                    "Result": result_str,
                    "Value": _fmt(getattr(line, "value", None), ".3f")
                    if getattr(line, "value", None) is not None
                    else "",
                    "Limit": _fmt(getattr(line, "limit", None), ".3f")
                    if getattr(line, "limit", None) is not None
                    else "",
                    "Margin": _fmt(getattr(line, "margin", None), ".3f")
                    if getattr(line, "margin", None) is not None
                    else "",
                    "Message": getattr(line, "message", "") or "",
                }
            )
        crit_df = pd.DataFrame(crit_rows)

    # --- Build Weight Items sheet rows (similar to PDF Weight Items table) ---
    items_rows = _build_weight_items_rows(ship, condition, results)
    df_items = pd.DataFrame(items_rows) if items_rows else None

    # --- Write all sheets and apply styling ---
    with pd.ExcelWriter(str(filepath), engine="openpyxl") as writer:
        # Sheet 1 – condition summary
        df_summary.to_excel(writer, sheet_name="Condition Summary", index=False)
        ws_summary = writer.sheets["Condition Summary"]
        ws_summary.column_dimensions["A"].width = 32
        ws_summary.column_dimensions["B"].width = 40
        _style_header(ws_summary)
        _style_body_table(ws_summary, start_row=2, first_col_bold=True, stripe=True)
        ws_summary.freeze_panes = "A2"

        # Sheet 2 – equilibrium data
        df_eq.to_excel(writer, sheet_name="Equilibrium Data", index=False)
        ws_eq = writer.sheets["Equilibrium Data"]
        ws_eq.column_dimensions["A"].width = 32
        ws_eq.column_dimensions["B"].width = 18
        ws_eq.column_dimensions["C"].width = 10
        _style_header(ws_eq)
        _style_body_table(ws_eq, start_row=2, first_col_bold=True, stripe=True)
        ws_eq.freeze_panes = "A2"

        # Sheet 3 – Weight items & FSM (optional, if we have any items)
        if df_items is not None and not df_items.empty:
            df_items.to_excel(writer, sheet_name="Weight Items", index=False)
            ws_items = writer.sheets["Weight Items"]
            # Column widths tuned to keep sheet readable and similar to PDF layout
            ws_items.column_dimensions["A"].width = 18  # Item
            ws_items.column_dimensions["B"].width = 14  # Quantity / Fill
            ws_items.column_dimensions["C"].width = 14  # Unit mass
            ws_items.column_dimensions["D"].width = 16  # Total mass
            ws_items.column_dimensions["E"].width = 16  # Long. arm
            ws_items.column_dimensions["F"].width = 16  # Vert. arm
            ws_items.column_dimensions["G"].width = 18  # Total FSM
            ws_items.column_dimensions["H"].width = 20  # FSM Type
            _style_header(ws_items)
            _style_body_table(ws_items, start_row=2, first_col_bold=True, stripe=True)
            ws_items.freeze_panes = "A2"

        # Sheet 4 – IMO / ancillary criteria (optional)
        if crit_df is not None and not crit_df.empty:
            crit_df.to_excel(writer, sheet_name="IMO Criteria", index=False)
            ws_crit = writer.sheets["IMO Criteria"]
            # Set reasonable widths
            ws_crit.column_dimensions["A"].width = 10  # Group
            ws_crit.column_dimensions["B"].width = 12  # Code
            ws_crit.column_dimensions["C"].width = 28  # Name
            ws_crit.column_dimensions["D"].width = 16  # Reference
            ws_crit.column_dimensions["E"].width = 10  # Result
            ws_crit.column_dimensions["F"].width = 14  # Value
            ws_crit.column_dimensions["G"].width = 14  # Limit
            ws_crit.column_dimensions["H"].width = 14  # Margin
            ws_crit.column_dimensions["I"].width = 50  # Message
            _style_header(ws_crit)
            # Body styling: zebra stripes, wrap long text in Name/Message
            _style_body_table(ws_crit, start_row=2, first_col_bold=False, stripe=True)
            ws_crit.freeze_panes = "A2"

            # Colour the Result column similar to the manual (green PASS, red FAIL, grey N/A).
            result_col_idx = list(crit_df.columns).index("Result") + 1
            for row_idx in range(2, ws_crit.max_row + 1):
                cell = ws_crit.cell(row=row_idx, column=result_col_idx)
                value = (str(cell.value) or "").upper()
                if "PASS" in value:
                    cell.fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
                elif "FAIL" in value:
                    cell.fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
                elif "N_A" in value or "N/A" in value:
                    cell.fill = PatternFill(fill_type="solid", fgColor="E7E6E6")

                # Center-align Result values for a cleaner look
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Wrap text for Name and Message columns so long text fits neatly
            name_col_idx = list(crit_df.columns).index("Name") + 1
            msg_col_idx = list(crit_df.columns).index("Message") + 1
            for row_idx in range(2, ws_crit.max_row + 1):
                for col_idx in (name_col_idx, msg_col_idx):
                    c = ws_crit.cell(row=row_idx, column=col_idx)
                    c.alignment = Alignment(
                        horizontal="left",
                        vertical="top",
                        wrap_text=True,
                    )

            # Rotate header labels vertically so long words fit in narrower columns
            for cell in ws_crit[1]:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="bottom",
                    text_rotation=90,
                    wrap_text=True,
                )
